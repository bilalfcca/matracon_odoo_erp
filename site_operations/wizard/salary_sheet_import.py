"""Salary Sheet Import Wizard.

Parses an Excel (.xlsx) or CSV file and creates / replaces the payroll
lines on a draft salary sheet.

Expected columns (case-insensitive, spaces/underscores stripped):
  • Employee (required)   — matched by name (case-insensitive partial match)
  • Paid Days (required)
  • Basic Salary          — leave blank to auto-compute from employee record
  • Allowances            — leave blank to auto-compute
  • WHT                   — leave blank to auto-compute
  • EOBI                  — leave blank to auto-compute
  • Advance to Deduct     — leave blank to auto-compute
  • Backcharge to Deduct  — leave blank to auto-compute

If an override cell is blank, the value is recalculated from the
employee's current settings (same logic as manual entry onchange).
"""
import base64
import calendar
import csv
import io

from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class SalarySheetImport(models.TransientModel):
    _name = 'x.salary.sheet.import'
    _description = 'Import Salary Sheet Lines'

    sheet_id = fields.Many2one(
        'x.salary.sheet', required=True,
        default=lambda self: self.env.context.get('default_sheet_id'),
    )
    import_file = fields.Binary(
        string='File (Excel or CSV)', required=True,
        help='Upload an .xlsx or .csv file. '
             'Download the template first to see the required columns.',
    )
    import_filename = fields.Char(string='Filename')
    replace_existing = fields.Boolean(
        string='Replace All Existing Lines', default=True,
        help='Checked: removes all current payroll lines before importing.\n'
             'Unchecked: adds new employees; updates paid-days of existing ones.',
    )
    preview_info = fields.Char(
        string='Last Import Result', readonly=True,
    )

    # ─────────────────────────────────────────────────────────────────────
    # COLUMN NORMALISER
    # ─────────────────────────────────────────────────────────────────────

    _HEADER_MAP = {
        'employee':           'employee',
        'employeename':       'employee',
        'name':               'employee',
        'staff':              'employee',
        'paiddays':           'paid_days',
        'days':               'paid_days',
        'workingdays':        'paid_days',
        'presentdays':        'paid_days',
        'basic':              'basic_salary',
        'basicsalary':        'basic_salary',
        'basicpaiddays':      'basic_salary',
        'allowances':         'allowances',
        'totalallowances':    'allowances',
        'wht':                'wht',
        'tax':                'wht',
        'incometax':          'wht',
        'withholdingtax':     'wht',
        'eobi':               'eobi',
        'advance':            'advance_to_deduct',
        'advancetodeduct':    'advance_to_deduct',
        'advancerecovery':    'advance_to_deduct',
        'backcharge':         'backcharge_to_deduct',
        'backchargetodeduct': 'backcharge_to_deduct',
        'backchargerecovery': 'backcharge_to_deduct',
    }

    @api.model
    def _norm(self, header_raw):
        """Normalise a raw column header to a canonical key."""
        h = str(header_raw or '').lower().strip()
        h = h.replace(' ', '').replace('_', '').replace('-', '')
        return self._HEADER_MAP.get(h, h)

    # ─────────────────────────────────────────────────────────────────────
    # FILE PARSERS
    # ─────────────────────────────────────────────────────────────────────

    def _parse_excel(self, file_data):
        if not openpyxl:
            raise UserError(_(
                'openpyxl is not installed on this server. '
                'Please upload a CSV file instead, or contact your administrator.'))
        wb = openpyxl.load_workbook(
            io.BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        header = [self._norm(h) for h in rows[0]]
        result = []
        for raw_row in rows[1:]:
            if not any(c for c in raw_row if c is not None):
                continue
            row = {header[i]: (raw_row[i] if i < len(raw_row) else '')
                   for i in range(len(header))}
            result.append(row)
        return result

    def _parse_csv(self, file_data):
        try:
            text = file_data.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = file_data.decode('latin-1')
        reader = csv.DictReader(io.StringIO(text))
        result = []
        for raw_row in reader:
            if not any(str(v).strip() for v in raw_row.values()):
                continue
            result.append({self._norm(k): str(v).strip()
                           for k, v in raw_row.items()})
        return result

    # ─────────────────────────────────────────────────────────────────────
    # HELPERS
    # ─────────────────────────────────────────────────────────────────────

    @staticmethod
    def _to_float(val):
        """Convert a cell value to float; returns None if blank/invalid."""
        if val is None or str(val).strip() == '':
            return None
        try:
            return float(str(val).replace(',', ''))
        except (ValueError, TypeError):
            return None

    # ─────────────────────────────────────────────────────────────────────
    # MAIN ACTION — IMPORT
    # ─────────────────────────────────────────────────────────────────────

    def action_import(self):
        self.ensure_one()
        sheet = self.sheet_id
        if sheet.state != 'draft':
            raise UserError(_('Only draft salary sheets can have lines imported.'))
        if not self.import_file:
            raise UserError(_('Please upload a file before importing.'))

        filename = (self.import_filename or '').lower()
        file_data = base64.b64decode(self.import_file)

        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            rows = self._parse_excel(file_data)
        elif filename.endswith('.csv'):
            rows = self._parse_csv(file_data)
        else:
            raise UserError(_(
                'Unsupported file type "%s". Please upload an .xlsx or .csv file.')
                % (self.import_filename or ''))

        if not rows:
            raise UserError(_('The file appears to be empty or has no data rows.'))

        # Validate required columns
        first = rows[0]
        if 'employee' not in first:
            raise UserError(_(
                'Could not find an "Employee" column in the file.\n'
                'Please download the template and use the exact column headers.'))
        if 'paid_days' not in first:
            raise UserError(_(
                'Could not find a "Paid Days" column in the file.\n'
                'Please download the template and use the exact column headers.'))

        Employee = self.env['hr.employee'].sudo()
        Backcharge = self.env['x.employee.backcharge'].sudo()
        SalaryLine = self.env['x.salary.sheet.line'].sudo()

        days_in_month = 30
        if sheet.date_from:
            days_in_month = calendar.monthrange(
                sheet.date_from.year, sheet.date_from.month)[1]

        parsed = []
        errors = []

        for i, row in enumerate(rows, start=2):
            emp_name = str(row.get('employee', '') or '').strip()
            if not emp_name:
                continue  # silently skip blank employee rows

            paid_days_raw = row.get('paid_days', '')
            paid_days = self._to_float(paid_days_raw)
            if paid_days is None:
                errors.append(
                    _('Row %(row)d: invalid Paid Days value "%(val)s" for employee "%(emp)s".')
                    % {'row': i, 'val': paid_days_raw, 'emp': emp_name})
                continue
            if paid_days < 0 or paid_days > days_in_month + 1:
                errors.append(
                    _('Row %(row)d: Paid Days %(days)s is out of range for employee "%(emp)s".')
                    % {'row': i, 'days': paid_days, 'emp': emp_name})
                continue

            emp = Employee.search([('name', 'ilike', emp_name)], limit=1)
            if not emp:
                errors.append(
                    _('Row %(row)d: Employee "%(emp)s" not found in the system.')
                    % {'row': i, 'emp': emp_name})
                continue

            parsed.append((i, emp, paid_days, row))

        if errors:
            raise UserError(
                _('The following errors must be corrected before importing:\n\n')
                + '\n'.join(f'• {e}' for e in errors))

        if not parsed:
            raise UserError(_('No valid employee rows found in the file.'))

        # ── Replace or update ────────────────────────────────────────────
        if self.replace_existing:
            sheet.line_ids.sudo().unlink()
            existing_map = {}
        else:
            existing_map = {
                line.employee_id.id: line
                for line in sheet.line_ids
            }

        created = updated = 0

        for _row_num, emp, paid_days, row in parsed:
            factor = paid_days / days_in_month if days_in_month else 0.0

            basic_ov     = self._to_float(row.get('basic_salary'))
            allow_ov     = self._to_float(row.get('allowances'))
            wht_ov       = self._to_float(row.get('wht'))
            eobi_ov      = self._to_float(row.get('eobi'))
            advance_ov   = self._to_float(row.get('advance_to_deduct'))
            backcharge_ov = self._to_float(row.get('backcharge_to_deduct'))

            # Auto-compute from employee when override is blank
            basic = basic_ov if basic_ov is not None else (emp.x_basic_salary or 0.0) * factor
            hra   = (emp.x_hra or 0.0) * factor
            site_allow = (emp.x_site_allowance or 0.0) * factor
            allowances = allow_ov if allow_ov is not None else hra + site_allow
            gross = basic + allowances

            wht  = wht_ov  if wht_ov  is not None else gross * (emp.x_wht_rate or 0.0) / 100.0
            eobi = eobi_ov if eobi_ov is not None else (emp.x_eobi_amount or 0.0)

            emp_advance = emp.x_advance_balance or 0.0
            advance = advance_ov if advance_ov is not None else min(emp_advance, gross)

            bcs = Backcharge.search([
                ('employee_id', '=', emp.id),
                ('state', 'in', ('open', 'partial')),
            ])
            emp_backcharge = sum(bcs.mapped('remaining_amount'))
            backcharge = backcharge_ov if backcharge_ov is not None else emp_backcharge

            deductions = wht + eobi + advance + backcharge
            net = max(gross - deductions, 0.0)

            vals = {
                'sheet_id': sheet.id,
                'employee_id': emp.id,
                'paid_days': paid_days,
                'basic_salary':          round(basic, 2),
                'allowances':            round(allowances, 2),
                'deductions':            round(deductions, 2),
                'net_payable':           round(net, 2),
                'detail_hra':            round(hra, 2),
                'detail_site_allowance': round(site_allow, 2),
                'detail_wht':            round(wht, 2),
                'detail_eobi':           round(eobi, 2),
                'detail_advance':        round(advance, 2),
                'outstanding_advance':   round(emp_advance, 2),
                'advance_to_deduct':     round(advance, 2),
                'outstanding_backcharge': round(emp_backcharge, 2),
                'backcharge_to_deduct':  round(backcharge, 2),
            }

            existing = existing_map.get(emp.id)
            if existing:
                existing.write(vals)
                updated += 1
            else:
                SalaryLine.create(vals)
                created += 1

        sheet.message_post(
            body=_(
                'Salary lines imported from file <b>%(filename)s</b>: '
                '%(created)d created, %(updated)d updated.'
            ) % {
                'filename': self.import_filename or 'file',
                'created': created,
                'updated': updated,
            }
        )

        # Return to the salary sheet form
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'x.salary.sheet',
            'view_mode': 'form',
            'res_id': sheet.id,
            'target': 'current',
        }

    # ─────────────────────────────────────────────────────────────────────
    # DOWNLOAD TEMPLATE
    # ─────────────────────────────────────────────────────────────────────

    def action_download_template(self):
        """Generate and return an Excel import template."""
        if not openpyxl:
            raise UserError(_('openpyxl is not available. Please use a CSV file instead.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Salary Import'

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        headers = [
            ('Employee',           'Required — exact or partial name match', True),
            ('Paid Days',          'Required — number of paid days (decimals allowed)', True),
            ('Basic Salary',       'Leave blank to auto-compute from employee record', False),
            ('Allowances',         'Leave blank to auto-compute (HRA + Site Allowance)', False),
            ('WHT',                'Leave blank to auto-compute', False),
            ('EOBI',               'Leave blank to auto-compute', False),
            ('Advance to Deduct',  'Leave blank to auto-compute', False),
            ('Backcharge to Deduct', 'Leave blank to auto-compute', False),
        ]

        # Style header row
        req_fill  = PatternFill(fill_type='solid', fgColor='1E40AF')  # dark blue
        opt_fill  = PatternFill(fill_type='solid', fgColor='374151')  # dark grey
        white_font = Font(bold=True, color='FFFFFF')

        for col_idx, (col_name, comment, required) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = white_font
            cell.fill = req_fill if required else opt_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 4, 18)

        # Example data row
        ws.append(['Ahmad Ali', 26, '', '', '', '', '', ''])
        ws.append(['Fatima Malik', 22, '', '', '', '', '', ''])

        # Notes row
        ws.append([])
        notes_cell = ws.cell(row=5, column=1,
                             value='Blue columns are required. Grey columns are optional — '
                                   'leave blank to auto-compute from the employee master.')
        notes_cell.font = Font(italic=True, color='6B7280')
        ws.merge_cells('A5:H5')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        file_b64 = base64.b64encode(buf.read())

        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'salary_import_template.xlsx',
            'type': 'binary',
            'datas': file_b64,
            'mimetype':
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': 'x.salary.sheet.import',
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
