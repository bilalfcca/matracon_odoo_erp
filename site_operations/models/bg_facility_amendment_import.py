"""
Bank Facility Import Wizard  (x.bg.facility.import.wizard)
BG Amendment Import Wizard   (x.bg.amendment.import.wizard)

Both wizards follow the same pattern as bg_import_wizard.py:
  - Download Template button  → generates an XLSX with headers + example row
  - Upload File + Import      → reads the file and creates/updates records
"""
import base64
import datetime as dt
import logging
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from markupsafe import Markup

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


# ─── Shared helpers ───────────────────────────────────────────────────────────

def _parse_date(val):
    if val is None or val == '' or str(val).strip().upper() == 'N/A':
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    # Try common string formats
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(val, default=0.0):
    if val is None or str(val).strip() in ('', 'N/A'):
        return default
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return default


def _styled_header(ws, col, text, width=20):
    """Write a styled header cell."""
    cell = ws.cell(row=1, column=col, value=text)
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', fgColor='2E75B6')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFFFFF')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.column_dimensions[get_column_letter(col)].width = width


def _example_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill('solid', fgColor='EFF5FB')
    cell.alignment = Alignment(horizontal='left', vertical='center')


def _to_xlsx_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue())


# ─── Bank Facility Import Wizard ──────────────────────────────────────────────

class BgFacilityImportWizard(models.TransientModel):
    _name = 'x.bg.facility.import.wizard'
    _description = 'Bank Facility Import Wizard'

    excel_file     = fields.Binary('Excel File (.xlsx)', attachment=False)
    excel_filename = fields.Char()
    state          = fields.Selection([('draft', 'Ready'), ('done', 'Done')], default='draft')
    result_html    = fields.Html('Result', readonly=True)
    imported_count = fields.Integer(readonly=True)
    skipped_count  = fields.Integer(readonly=True)
    error_count    = fields.Integer(readonly=True)

    # ── Template download ─────────────────────────────────────────────────────
    def action_download_template(self):
        """Generate and return a ready-to-fill XLSX template."""
        self.ensure_one()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bank Facilities'
        ws.row_dimensions[1].height = 36

        headers = [
            ('Bank / Journal Name',   22),
            ('Reference No',          20),
            ('Date of Issuance\n(DD/MM/YYYY)', 22),
            ('Expiry Date\n(DD/MM/YYYY)',       22),
            ('Cash Margin %\n(e.g. 5.00)',      18),
            ('Commission %\n(e.g. 1.25)',        18),
            ('Sanctioned Limit',                 20),
        ]
        for col, (title, width) in enumerate(headers, 1):
            _styled_header(ws, col, title, width)

        # Example row
        ex = ['MCB Bank Limited', 'LTR/2024/001', '01/01/2024', '31/12/2025',
              '10.00', '1.25', '500000000']
        for col, val in enumerate(ex, 1):
            _example_cell(ws, 2, col, val)

        # Instructions sheet
        info = wb.create_sheet('Instructions')
        instructions = [
            ('Column',          'Notes'),
            ('Bank / Journal Name', 'Must exactly match a Bank journal in Odoo (Accounting → Journals → Type = Bank).'),
            ('Reference No',        'Sanction letter / facility reference number from the bank.'),
            ('Date of Issuance',    'Date the facility letter was issued. Format: DD/MM/YYYY.'),
            ('Expiry Date',         'Facility expiry date. Format: DD/MM/YYYY.'),
            ('Cash Margin %',       'Flat cash margin % applied to all BGs under this facility (e.g. 10.00 for 10%).'),
            ('Commission %',        'Bank commission / pricing % (e.g. 1.25 for 1.25%).'),
            ('Sanctioned Limit',    'Total limit amount sanctioned by the bank (numeric, no commas required).'),
            ('',                    ''),
            ('Notes',               'Row 1 = headers (do not delete). Row 2 = example (delete before importing).'),
            ('',                    'Rows where Bank / Journal Name is empty are ignored automatically.'),
            ('',                    'If a facility already exists for that bank, its Reference No and Sanctioned Limit are updated.'),
        ]
        for r, (a, b) in enumerate(instructions, 1):
            info.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
            info.cell(row=r, column=2, value=b)
        info.column_dimensions['A'].width = 25
        info.column_dimensions['B'].width = 80

        self.write({
            'excel_file':     _to_xlsx_bytes(wb),
            'excel_filename': 'Bank_Facility_Import_Template.xlsx',
        })
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }

    # ── Import ────────────────────────────────────────────────────────────────
    def action_import(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file first.'))

        try:
            wb = openpyxl.load_workbook(
                BytesIO(base64.b64decode(self.excel_file)), data_only=True)
        except Exception as e:
            raise UserError(_('Cannot open file: %s') % e)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Skip row 0 (headers)
        imported, skipped, errors = [], [], []

        for ri, row in enumerate(rows[1:], start=2):
            bank_name = str(row[0] or '').strip() if row[0] else ''
            if not bank_name:
                continue   # blank row

            ref_no    = str(row[1] or '').strip() or False
            iss_date  = _parse_date(row[2])
            exp_date  = _parse_date(row[3])
            margin_pct = _parse_float(row[4])
            comm_pct   = _parse_float(row[5])
            total_limit = _parse_float(row[6])

            # Find journal
            journal = self.env['account.journal'].sudo().search(
                [('type', '=', 'bank'), ('name', 'ilike', bank_name)], limit=1)
            if not journal:
                errors.append(
                    f'Row {ri}: Bank <b>"{bank_name}"</b> not found — '
                    f'create it in Accounting → Configuration → Journals first.')
                continue

            # Create or update facility
            facility = self.env['x.bank.guarantee.facility'].sudo().search(
                [('journal_id', '=', journal.id)], limit=1)

            vals = {
                'journal_id':         journal.id,
                'reference_no':       ref_no,
                'date_of_issuance':   iss_date,
                'expiry_date':        exp_date,
                'cash_margin_percent': margin_pct,
                'commission_percent': comm_pct,
                'total_limit':        total_limit,
            }

            if facility:
                facility.sudo().write(vals)
                skipped.append(f'Row {ri}: Updated existing facility for <b>{bank_name}</b>.')
            else:
                self.env['x.bank.guarantee.facility'].sudo().create(vals)
                imported.append(f'Row {ri}: Created facility for <b>{bank_name}</b>.')

        # Build result
        parts = []
        if imported:
            rows_html = ''.join(f'<li>{s}</li>' for s in imported)
            parts.append(
                f'<div class="alert alert-success mb-2">'
                f'<b>✓ {len(imported)} facility(ies) created.</b><ul>{rows_html}</ul></div>')
        if skipped:
            rows_html = ''.join(f'<li>{s}</li>' for s in skipped)
            parts.append(
                f'<div class="alert alert-info mb-2">'
                f'<b>{len(skipped)} facility(ies) updated.</b><ul>{rows_html}</ul></div>')
        if errors:
            rows_html = ''.join(f'<li>{e}</li>' for e in errors)
            parts.append(
                f'<div class="alert alert-danger mb-2">'
                f'<b>{len(errors)} error(s):</b><ul>{rows_html}</ul></div>')
        if not parts:
            parts.append('<div class="alert alert-warning">No data rows found in the file.</div>')

        self.write({
            'state':          'done',
            'result_html':    Markup('\n'.join(parts)),
            'imported_count': len(imported),
            'skipped_count':  len(skipped),
            'error_count':    len(errors),
        })
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }

    def action_view_facilities(self):
        return {
            'type':      'ir.actions.act_window',
            'name':      _('Bank Facility Limits'),
            'res_model': 'x.bank.guarantee.facility',
            'view_mode': 'list,form',
        }


# ─── BG Amendment Import Wizard ───────────────────────────────────────────────

_AMEND_TYPE_MAP = {
    'ORIGINAL':  'initial',
    'INITIAL':   'initial',
    'EXTENSION': 'extension',
    'INCREASE':  'increase',
    'DECREASE':  'decrease',
    'RELEASE':   'release',
    'OTHER':     'other',
}


class BgAmendmentImportWizard(models.TransientModel):
    _name = 'x.bg.amendment.import.wizard'
    _description = 'BG Amendment Import Wizard'

    excel_file     = fields.Binary('Excel File (.xlsx)', attachment=False)
    excel_filename = fields.Char()
    state          = fields.Selection([('draft', 'Ready'), ('done', 'Done')], default='draft')
    result_html    = fields.Html('Result', readonly=True)
    imported_count = fields.Integer(readonly=True)
    skipped_count  = fields.Integer(readonly=True)
    error_count    = fields.Integer(readonly=True)

    # ── Template download ─────────────────────────────────────────────────────
    def action_download_template(self):
        self.ensure_one()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'BG Amendments'
        ws.row_dimensions[1].height = 36

        headers = [
            ('Guarantee No\n(must exist in Odoo)',    26),
            ('Amendment Date\n(DD/MM/YYYY)',           22),
            ('Amendment Type\n(see Instructions)',     24),
            ('Description\n(required)',                40),
            ('Amount Change\n(+ve or -ve)',            20),
            ('New Expiry Date\n(DD/MM/YYYY, optional)', 24),
        ]
        for col, (title, width) in enumerate(headers, 1):
            _styled_header(ws, col, title, width)

        # Two example rows
        examples = [
            ['BG/2024/0001', '15/03/2025', 'Extension', 'Extension of expiry by 6 months', 0, '30/09/2025'],
            ['BG/2024/0002', '10/04/2025', 'Increase',  'Increased BG amount per variation order #3', 5000000, ''],
        ]
        for ri, ex in enumerate(examples, 2):
            for ci, val in enumerate(ex, 1):
                _example_cell(ws, ri, ci, val)

        # Instructions sheet
        info = wb.create_sheet('Instructions')
        instructions = [
            ('Column',           'Notes'),
            ('Guarantee No',     'Must match exactly the Guarantee No. printed on the BG in Odoo (e.g. BG-MCB-2024-001).'),
            ('Amendment Date',   'Date of this amendment. Format: DD/MM/YYYY.'),
            ('Amendment Type',   'One of: Original, Extension, Increase, Decrease, Release, Other (case-insensitive).'),
            ('Description',      'Required. Brief description of the amendment.'),
            ('Amount Change',    'Positive for increases, negative for decreases, 0 for extensions/release.'),
            ('New Expiry Date',  'Fill only for Extension type. Leave blank for others. Format: DD/MM/YYYY.'),
            ('', ''),
            ('Notes',            'Row 1 = headers. Rows 2-3 = examples (delete before importing).'),
            ('',                 'Amendments are ADDED to the matching BG — duplicates are not checked.'),
            ('',                 'For Extension type: the parent BG expiry date is automatically updated to New Expiry Date.'),
        ]
        for r, (a, b) in enumerate(instructions, 1):
            info.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
            info.cell(row=r, column=2, value=b)
        info.column_dimensions['A'].width = 22
        info.column_dimensions['B'].width = 90

        self.write({
            'excel_file':     _to_xlsx_bytes(wb),
            'excel_filename': 'BG_Amendment_Import_Template.xlsx',
        })
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }

    # ── Import ────────────────────────────────────────────────────────────────
    def action_import(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file first.'))

        try:
            wb = openpyxl.load_workbook(
                BytesIO(base64.b64decode(self.excel_file)), data_only=True)
        except Exception as e:
            raise UserError(_('Cannot open file: %s') % e)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Cache BGs by guarantee_number for fast lookup
        bg_cache = {}
        imported, skipped, errors = [], [], []
        today = fields.Date.context_today(self)

        for ri, row in enumerate(rows[1:], start=2):
            gno = str(row[0] or '').strip() if row[0] else ''
            if not gno:
                continue  # blank row

            amend_date   = _parse_date(row[1]) or today
            amend_type_raw = str(row[2] or '').strip().upper()
            amend_type   = _AMEND_TYPE_MAP.get(amend_type_raw, 'other')
            description  = str(row[3] or '').strip()
            amount_change = _parse_float(row[4])
            new_expiry   = _parse_date(row[5]) if len(row) > 5 else None

            if not description:
                errors.append(f'Row {ri} ({gno}): Description is required — skipped.')
                continue

            # Find parent BG
            if gno not in bg_cache:
                bg = self.env['x.bank.guarantee'].sudo().search(
                    [('guarantee_number', '=', gno)], limit=1)
                bg_cache[gno] = bg
            bg = bg_cache[gno]

            if not bg:
                errors.append(
                    f'Row {ri}: Guarantee No <b>"{gno}"</b> not found in Odoo — skipped.')
                continue

            vals = {
                'guarantee_id':   bg.id,
                'amendment_date': amend_date,
                'amendment_type': amend_type,
                'description':    description,
                'amount_change':  amount_change,
                'new_expiry_date': new_expiry or False,
            }

            try:
                self.env['x.bank.guarantee.amendment'].sudo().create(vals)
                imported.append(
                    f'Row {ri}: Amendment added to <b>{gno}</b> '
                    f'({amend_type}, {amend_date}).')
            except Exception as exc:
                errors.append(f'Row {ri} ({gno}): ERROR — {exc}')
                _logger.exception('Amendment import error row %s (%s)', ri, gno)

        # Build result
        parts = []
        if imported:
            rows_html = ''.join(f'<li>{s}</li>' for s in imported)
            parts.append(
                f'<div class="alert alert-success mb-2">'
                f'<b>✓ {len(imported)} amendment(s) imported.</b><ul>{rows_html}</ul></div>')
        if skipped:
            rows_html = ''.join(f'<li>{s}</li>' for s in skipped)
            parts.append(
                f'<div class="alert alert-warning mb-2">'
                f'<b>{len(skipped)} skipped.</b><ul>{rows_html}</ul></div>')
        if errors:
            rows_html = ''.join(f'<li>{e}</li>' for e in errors)
            parts.append(
                f'<div class="alert alert-danger mb-2">'
                f'<b>{len(errors)} error(s):</b><ul>{rows_html}</ul></div>')
        if not parts:
            parts.append('<div class="alert alert-warning">No data rows found in the file.</div>')

        self.write({
            'state':          'done',
            'result_html':    Markup('\n'.join(parts)),
            'imported_count': len(imported),
            'skipped_count':  len(skipped),
            'error_count':    len(errors),
        })
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }

    def action_view_amendments(self):
        return {
            'type':      'ir.actions.act_window',
            'name':      _('BG Amendments'),
            'res_model': 'x.bank.guarantee.amendment',
            'view_mode': 'list,form',
        }
